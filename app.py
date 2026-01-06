from flask import Flask, request, jsonify, send_from_directory, render_template, send_file
from flask_cors import CORS
import os
import bcrypt
import jwt
import datetime
import pytz
import sys
import time
import threading
from functools import wraps

import sqlite3
from dotenv import load_dotenv
import openpyxl
from io import BytesIO

try:
    import pymssql
except ImportError:
    pymssql = None

load_dotenv()

app = Flask(__name__, static_folder='netlify', template_folder='netlify')
_secret = os.getenv('SECRET_KEY')
if not _secret:
    _secret = os.getenv('DB_PASSWORD', 'ponto_sre_carapina')
app.config['SECRET_KEY'] = _secret
try:
    # Allow all origins, methods, and headers for the API to support external hosting (Vercel/GitHub Pages)
    CORS(app, resources={r"/api/*": {"origins": "*"}}, 
         supports_credentials=True,
         allow_headers=["*"],
         methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])
except Exception:
    pass

# Routes to serve frontend files locally
@app.route('/')
def serve_index():
    return send_from_directory('netlify', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    if not path.startswith('api/'):
        return send_from_directory('netlify', path)
    return jsonify({'message': 'Not Found'}), 404

@app.before_request
def log_request_info():
    if request.path.startswith('/api/'):
        print(f"DEBUG REQ: {request.method} {request.path} | Origin: {request.headers.get('Origin')} | Remote: {request.remote_addr}")

# Database Configuration
server = os.getenv('DB_SERVER')
database = os.getenv('DB_NAME')
username = os.getenv('DB_USER')
password = os.getenv('DB_PASSWORD')
sqlite_path = os.getenv('SQLITE_PATH', 'local.db')
USE_SQLITE = False

# Optional: initialize database/tables on startup when deploying
if os.getenv('INIT_DB_ON_START', 'false').lower() == 'true':
    try:
        from setup_db import create_database, create_tables
        create_database()
        create_tables()
    except Exception:
        pass

# Global DB Status
DB_ONLINE = False

def check_db_status():
    global DB_ONLINE
    while True:
        try:
            conn = None
            success = False
            
            # Try pymssql first
            if pymssql:
                try:
                    conn = pymssql.connect(
                        server=server, user=username, password=password, database=database, 
                        login_timeout=3, autocommit=True
                    )
                    success = True
                except:
                    pass
            
            # Try pyodbc if pymssql failed/not present
            if not success:
                try:
                    import pyodbc
                    drivers = [d for d in pyodbc.drivers() if 'SQL Server' in d]
                    driver = drivers[0] if drivers else 'ODBC Driver 17 for SQL Server'
                    conn_str = f'DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};TrustServerCertificate=yes;Connection Timeout=3'
                    conn = pyodbc.connect(conn_str, autocommit=True)
                    success = True
                except:
                    pass
            
            if success:
                try:
                   if conn: conn.close()
                except: pass
                
                # Check for transition from Offline -> Online
                if not DB_ONLINE:
                    print("DEBUG: SQL Server connection restored. Triggering auto-sync.")
                    threading.Thread(target=auto_sync_all, daemon=True).start()
                
                DB_ONLINE = True
            else:
                DB_ONLINE = False
        except Exception:
            DB_ONLINE = False
            
        time.sleep(10)

def start_health_check():
    t = threading.Thread(target=check_db_status)
    t.daemon = True
    t.start()
    
def ensure_sqlite_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS Users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            role TEXT DEFAULT 'user'
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS TimeRecords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            matricula TEXT,
            user_name TEXT,
            record_type TEXT NOT NULL,
            timestamp DATETIME,
            neighborhood TEXT,
            city TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS OfflineQueue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            matricula TEXT,
            user_name TEXT,
            record_type TEXT NOT NULL,
            timestamp DATETIME,
            neighborhood TEXT,
            city TEXT
        )
    """)
    # Add columns if they don't exist
    try:
        c.execute("ALTER TABLE TimeRecords ADD COLUMN matricula TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE TimeRecords ADD COLUMN user_name TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE OfflineQueue ADD COLUMN matricula TEXT")
    except: pass
    try:
        c.execute("ALTER TABLE OfflineQueue ADD COLUMN user_name TEXT")
    except: pass
    conn.commit()

def migrate_local_data():
    """Populates matricula and user_name in existing TimeRecords and OfflineQueue entries."""
    print("DEBUG: Starting local data migration...")
    try:
        conn = sqlite3.connect(sqlite_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        # Build mapping
        cur.execute("SELECT id, matricula, name FROM Users")
        user_map = {r['id']: (r['matricula'], r['name']) for r in cur.fetchall()}
        
        # Update TimeRecords
        cur.execute("SELECT id, user_id FROM TimeRecords WHERE matricula IS NULL")
        to_update = cur.fetchall()
        for r in to_update:
            if r['user_id'] in user_map:
                m, n = user_map[r['user_id']]
                cur.execute("UPDATE TimeRecords SET matricula = ?, user_name = ? WHERE id = ?", (m, n, r['id']))
        
        # Update OfflineQueue
        cur.execute("SELECT id, user_id FROM OfflineQueue WHERE matricula IS NULL")
        to_update_q = cur.fetchall()
        for r in to_update_q:
            if r['user_id'] in user_map:
                m, n = user_map[r['user_id']]
                cur.execute("UPDATE OfflineQueue SET matricula = ?, user_name = ? WHERE id = ?", (m, n, r['id']))
        
        conn.commit()
        conn.close()
        print("DEBUG: Local data migration complete.")
    except Exception as e:
        print(f"DEBUG: Migration error: {e}")
    # The original instruction had conn.commit() here, but it should be inside the try block before conn.close()
    # or removed if the previous commit covers it. Given the structure, it's likely a copy-paste error.
    # I'll remove the redundant conn.commit() here as it's already done inside the try block.

def ensure_default_admin():
    try:
        admin_mat = os.getenv('ADMIN_MATRICULA', 'admin')
        admin_pass = os.getenv('ADMIN_PASSWORD', 'admin')
        admin_name = os.getenv('ADMIN_NAME', 'Administrador')
        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        ensure_sqlite_schema(sconn)
        scur = sconn.cursor()
        scur.execute("SELECT 1 FROM Users WHERE matricula = ?", (admin_mat,))
        exists = scur.fetchone()
        if not exists:
            hashed = bcrypt.hashpw(admin_pass.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            try:
                scur.execute("INSERT OR IGNORE INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)", (admin_mat, hashed, admin_name, 'admin'))
                sconn.commit()
            except Exception:
                pass
        sconn.close()
    except Exception:
        pass

def get_db_connection():
    if DB_ONLINE:
        # Attempt SQL Server (pymssql)
        try:
            if pymssql:
                conn = pymssql.connect(
                    server=server, user=username, password=password, database=database, 
                    as_dict=True, autocommit=True, login_timeout=3
                )
                try:
                    with conn.cursor() as cur:
                        cur.execute("SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED")
                except:
                    pass
                return conn
        except Exception:
            pass

        # Attempt SQL Server (pyodbc)
        try:
            import pyodbc
            drivers = [d for d in pyodbc.drivers() if 'SQL Server' in d]
            driver = drivers[0] if drivers else 'ODBC Driver 17 for SQL Server'
            conn_str = f'DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};TrustServerCertificate=yes;Connection Timeout=3'
            conn = pyodbc.connect(conn_str, autocommit=True)
            try:
                cur = conn.cursor()
                cur.execute("SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED")
            except:
                pass
            return conn
        except Exception:
            pass

    # Fallback to SQLite
    conn = sqlite3.connect(sqlite_path)
    conn.row_factory = sqlite3.Row
    ensure_sqlite_schema(conn)
    return conn


def sql_online():
    return DB_ONLINE

def get_ph(conn):
    """Returns the correct SQL placeholder based on the connection type."""
    if isinstance(conn, sqlite3.Connection):
        return '?'
    if 'pymssql' in str(type(conn)):
        return '%s'
    return '?'

def get_user_info_by_id(user_id, conn):
    """Returns (matricula, name) for a given user_id using the provided connection."""
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cur = conn.cursor()
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cur.execute(f"SELECT matricula, name FROM Users {nolock} WHERE id = {ph}", (user_id,))
        row = cur.fetchone()
        if row:
            return rf(row, 'matricula'), rf(row, 'name')
    except:
        pass
    return None, None

def get_user_info_by_matricula(matricula, conn):
    """Returns (id, name) for a given matricula using the provided connection."""
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cur = conn.cursor()
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cur.execute(f"SELECT id, name FROM Users {nolock} WHERE matricula = {ph}", (matricula,))
        row = cur.fetchone()
        if row:
            return rf(row, 'id'), rf(row, 'name')
    except:
        pass
    return None, None


def rf(row, name):
    try:
        if isinstance(row, dict):
            return row.get(name)
        return row[name]
    except Exception:
        try:
            return getattr(row, name)
        except AttributeError:
            return None

# Auth Decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "") or ""
        parts = auth.split()
        token = parts[1] if len(parts) == 2 and parts[0].lower() == "bearer" else None
        if not token:
            return jsonify({"message": "Token is missing!"}), 401
        try:
            data = jwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
            # Favor matricula for cross-system stability
            curr_user_mat = data.get("matricula")
            if not curr_user_mat:
                # Fallback for old tokens if any
                curr_user_mat = str(data.get("user_id"))
            role = data["role"]
        except Exception:
            return jsonify({"message": "Token is invalid!"}), 401
        return f(curr_user_mat, role, *args, **kwargs)
    return decorated

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200

@app.route('/register')
def register_page():
    return render_template('register.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/admin')
def admin_page():
    return render_template('admin.html')

# API Endpoints
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    
    try:
        query = f"INSERT INTO Users (matricula, password, name, role) VALUES ({ph}, {ph}, {ph}, {ph})"
        cursor.execute(query, (data['matricula'], hashed_password, data['name'], 'user'))
        if isinstance(conn, sqlite3.Connection):
            conn.commit()
        elif ph == '?':
            conn.commit()
        # mirror to local sqlite for offline login
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            scur.execute("INSERT OR IGNORE INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)",
                         (data['matricula'], hashed_password, data['name'], 'user'))
            sconn.commit()
            sconn.close()
        except Exception:
            pass
        return jsonify({'message': 'User registered successfully!'}), 201
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'unique' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matricula already exists!'}), 409
        return jsonify({'message': msg}), 500
    finally:
        conn.close()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    conn = get_db_connection()
    ph = get_ph(conn)
    user = None
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cursor = conn.cursor()
        if not is_sqlite:
            query = f"SELECT id, matricula, password, name, role FROM Users WITH (NOLOCK) WHERE matricula = {ph}"
        else:
            query = f"SELECT id, matricula, password, name, role FROM Users WHERE matricula = {ph}"
        cursor.execute(query, (data['matricula'],))
        user = cursor.fetchone()
        if user:
            print(f"DEBUG: Login - Matrícula {data['matricula']} encontrada no backend principal ({'SQLite' if is_sqlite else 'SQL Server'})")
        else:
            print(f"DEBUG: Login - Matrícula {data['matricula']} NÃO encontrada no backend principal ({'SQLite' if is_sqlite else 'SQL Server'})")
    except Exception as e:
        print(f"DEBUG: Login - Erro ao buscar usuario no backend principal: {e}")
        user = None
    finally:
        try:
            conn.close()
        except Exception:
            pass
    
    # Fallback to local sqlite if not found on primary connection (if primary was SQL)
    if not user and not is_sqlite:
        try:
            print(f"DEBUG: Login - Tentando fallback local para matrícula {data['matricula']}")
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            scur.execute("SELECT id, matricula, password, name, role FROM Users WHERE matricula = ?", (data['matricula'],))
            user = scur.fetchone()
            sconn.close()
            if user:
                print(f"DEBUG: Login - Matrícula {data['matricula']} encontrada no fallback local.")
            else:
                print(f"DEBUG: Login - Matrícula {data['matricula']} NÃO encontrada no fallback local.")
        except Exception as e:
            print(f"DEBUG: Login - Erro no fallback local: {e}")
            user = None
    
    if user:
        try:
            input_pass = data['password'].encode('utf-8')
            hashed_pass = rf(user, 'password').encode('utf-8')
            if bcrypt.checkpw(input_pass, hashed_pass):
                print(f"DEBUG: Login - Sucesso para matrícula {data['matricula']}")
            else:
                print(f"DEBUG: Login - FALHA na senha para matrícula {data['matricula']}")
                return jsonify({'message': 'Invalid credentials!'}), 401
        except Exception as e:
            print(f"DEBUG: Login - Erro na verificação de senha: {e}")
            return jsonify({'message': 'Internal auth error'}), 500
    else:
        return jsonify({'message': 'User not found!'}), 401

    if user: # redundant check but safe for logic flow
        # Mirror user to local sqlite for future offline login
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            current_hash = rf(user, 'password')
            
            scur.execute("SELECT 1 FROM Users WHERE matricula = ?", (data['matricula'],))
            exists = scur.fetchone()
            if exists:
                 scur.execute("UPDATE Users SET password = ?, name = ?, role = ? WHERE matricula = ?", 
                              (current_hash, rf(user, 'name'), rf(user, 'role'), data['matricula']))
            else:
                scur.execute("INSERT INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)",
                             (data['matricula'], current_hash, rf(user, 'name'), rf(user, 'role')))
            sconn.commit()
            sconn.close()
        except Exception:
            pass

        try:
            token = jwt.encode({
                'matricula': rf(user, 'matricula'),
                'user_id': rf(user, 'id'), # keep for backward compat if needed
                'role': rf(user, 'role'),
                'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
            }, app.config['SECRET_KEY'], algorithm="HS256")
            
            # Trigger background sync for this user immediately
            threading.Thread(target=perform_sync_for_user, args=(rf(user, 'matricula'),), daemon=True).start()
            
            return jsonify({'token': token, 'role': rf(user, 'role'), 'name': rf(user, 'name')})
        except Exception as e:
            return jsonify({'message': f'Internal Server Error: {str(e)}'}), 500
    
    return jsonify({'message': 'Invalid credentials!'}), 401

@app.route('/api/punch', methods=['POST'])
@token_required
def punch(curr_user_mat, role):
    data = request.get_json()
    # Expecting: type, neighborhood, city, timestamp (optional)
    
    conn = get_db_connection()
    # Determine basic status
    is_sqlite = isinstance(conn, sqlite3.Connection)
    ph = get_ph(conn)

    # Use provided timestamp if available, else use current server time
    provided_ts = data.get('timestamp')
    if provided_ts:
        try:
            # Expected format from JS: YYYY-MM-DD HH:MM:SS
            current_time = datetime.datetime.strptime(provided_ts, '%Y-%m-%d %H:%M:%S')
        except Exception as e:
            print(f"Error parsing provided timestamp '{provided_ts}': {e}")
            current_time = datetime.datetime.now(pytz.timezone('America/Sao_Paulo')).replace(tzinfo=None)
    else:
        current_time = datetime.datetime.now(pytz.timezone('America/Sao_Paulo')).replace(tzinfo=None)
    
    # fetch denormalized user fields if available
    user_matricula = curr_user_mat
    sql_user_id, user_name = get_user_info_by_matricula(user_matricula, conn) # Use 'conn' for potential SQL Server
    
    # Also find local user_id to catch orphaned local records
    lconn = sqlite3.connect(sqlite_path)
    lconn.row_factory = sqlite3.Row
    local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
    lconn.close()

    if not user_name:
         user_name = l_user_name
    
    # Fallback to local user_id if SQL one not found (rare if online)
    sync_user_id = sql_user_id if sql_user_id else local_user_id
    
    # 1. Try Online Insert if applicable
    inserted_online = False
    
    if not is_sqlite:
        try:
            cursor = conn.cursor()
            # Insert into Online TimeRecords
            query = f"INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, neighborhood, city, timestamp) VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})"
            cursor.execute(query, (sync_user_id, user_matricula, user_name, data['type'], data.get('neighborhood'), data.get('city'), current_time))
            if ph == '?': # likely pyodbc
                conn.commit()
            inserted_online = True
        except Exception as e:
            print(f"Error inserting online: {e}")

    # 2. Local fallback if needed
    if is_sqlite or not inserted_online:
        # We need a dedicated sqlite connection for the queue to ensure we don't mix with a broken pymssql conn
        try:
            qconn = sqlite3.connect(sqlite_path)
            qconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(qconn)
            # Ensure we have user info for the local sqlite
            if not user_name:
                 user_id, user_name = get_user_info_by_matricula(user_matricula, qconn)
            
            qcur = qconn.cursor()
            try:
                qcur.execute(
                    """
                    INSERT INTO OfflineQueue (user_id, matricula, user_name, record_type, neighborhood, city, timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (sync_user_id, user_matricula, user_name, data['type'], data.get('neighborhood'), data.get('city'), current_time)
                )
                qconn.commit()
            finally:
                qconn.close()
        except Exception as e:
            # If even local save fails, then we return error
            return jsonify({'message': f'Error saving punch: {str(e)}'}), 500

    # Commit main connection if it was used and no autocommit (though we set autocommit=True for pymssql)
    if inserted_online and is_sqlite: 
         conn.commit()

    try:
        conn.close()
    except:
        pass

    return jsonify({'message': 'Ponto recorded successfully!'}), 201

@app.route('/api/history', methods=['GET'])
@token_required
def history(curr_user_mat, role):
    conn = get_db_connection()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        ph = get_ph(conn)
        
        user_matricula = curr_user_mat

        # Get user info for inclusive filtering
        sql_user_id, user_name = None, None
        local_user_id, l_user_name = None, None
        if sql_online():
            try:
                # Use a temp connection to avoid mixing with 'conn' if it's already used
                tconn = get_db_connection()
                sql_user_id, user_name = get_user_info_by_matricula(user_matricula, tconn)
                tconn.close()
            except: pass
        
        try:
            lconn = sqlite3.connect(sqlite_path)
            lconn.row_factory = sqlite3.Row
            local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
            lconn.close()
        except: pass

        records = []
        seen = set()
        
        if is_sqlite and sql_online():
            # Fallback detected but SQL is online - attempt forced SQL
            try:
                import pymssql
                fconn = pymssql.connect(server=server, user=username, password=password, database=database, as_dict=True, autocommit=True)
                try:
                    fcur = fconn.cursor()
                    fcur.execute(f"""
                        SELECT record_type, timestamp, neighborhood, city 
                        FROM TimeRecords WITH (NOLOCK)
                        WHERE matricula = {ph} 
                          AND MONTH(timestamp) = MONTH(GETDATE()) 
                          AND YEAR(timestamp) = YEAR(GETDATE())
                        ORDER BY timestamp DESC
                    """, (user_matricula,))
                    
                    sql_rows = fcur.fetchall()
                    for row in sql_rows:
                        ts = rf(row, 'timestamp')
                        if isinstance(ts, datetime.datetime):
                            ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                        key = (rf(row, 'record_type'), ts)
                        seen.add(key)
                        records.append({
                            'type': rf(row, 'record_type'),
                            'timestamp': ts,
                            'neighborhood': rf(row, 'neighborhood'),
                            'city': rf(row, 'city'),
                            'pending': False
                        })
                finally:
                    fconn.close()
            except Exception:
                pass
        
        if not records or is_sqlite: # Always check local if on SQLite fallback or if empty
            try:
                cursor = conn.cursor()
                if not is_sqlite:
                    # Query SQL Server by matricula
                    cursor.execute(f"""
                        SELECT record_type, timestamp, neighborhood, city 
                        FROM TimeRecords WITH (NOLOCK)
                        WHERE matricula = {ph} 
                          AND MONTH(timestamp) = MONTH(GETDATE()) 
                          AND YEAR(timestamp) = YEAR(GETDATE())
                        ORDER BY timestamp DESC
                    """, (user_matricula,))
                else:
                     # Query local SQLite by matricula
                     cursor.execute(f"""
                        SELECT record_type, timestamp, neighborhood, city 
                        FROM TimeRecords 
                        WHERE matricula = {ph} 
                          AND strftime('%m', timestamp) = strftime('%m', 'now') 
                          AND strftime('%Y', timestamp) = strftime('%Y', 'now')
                        ORDER BY timestamp DESC
                    """, (user_matricula,))
                
                rows = cursor.fetchall()
                for row in rows:
                    ts = rf(row, 'timestamp')
                    if isinstance(ts, datetime.datetime):
                        ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                    key = (rf(row, 'record_type'), ts)
                    if key not in seen:
                        seen.add(key)
                        records.append({
                            'type': rf(row, 'record_type'),
                            'timestamp': ts,
                            'neighborhood': rf(row, 'neighborhood'),
                            'city': rf(row, 'city'),
                            'pending': is_sqlite
                        })
            except Exception:
                pass

        # Append offline queued items by matricula or user_id
        try:
            sconn = sqlite3.connect(sqlite_path)
            try:
                sconn.row_factory = sqlite3.Row
                ensure_sqlite_schema(sconn)
                scur = sconn.cursor()
                scur.execute("SELECT record_type, timestamp, neighborhood, city FROM OfflineQueue WHERE matricula = ? OR (matricula IS NULL AND user_id = ?) OR (matricula IS NULL AND user_id = ?) ORDER BY timestamp DESC", (user_matricula, local_user_id, sql_user_id))
                qrows = scur.fetchall()
                for row in qrows:
                    ts = rf(row, 'timestamp')
                    if isinstance(ts, datetime.datetime):
                        ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                    records.append({
                        'type': rf(row, 'record_type'),
                        'timestamp': ts,
                        'neighborhood': rf(row, 'neighborhood'),
                        'city': rf(row, 'city'),
                        'pending': True
                    })
            finally:
                sconn.close()
        except Exception:
            pass
        
        return jsonify(records)
    finally:
        try:
            conn.close()
        except:
            pass

@app.route('/api/online')
def online():
    # Se o endpoint foi chamado, o servidor está online.
    # Verificamos o banco apenas para informação.
    db_ok = sql_online()
    return jsonify({'online': True, 'db_online': db_ok}), 200

@app.route('/api/user/report', methods=['GET'])
@token_required
def get_user_report(curr_user_mat, role):
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        user_matricula = curr_user_mat
        base = f"""
            SELECT t.matricula, t.user_name AS name,
                   t.record_type, t.timestamp, t.neighborhood, t.city
            FROM TimeRecords t
        """
        params = []
        
        # Add WHERE clause for matricula
        base += f" WHERE t.matricula = {ph}"
        params.append(user_matricula)

        if start_date:
            # Handle both SQL Server and SQLite for date filtering
            if isinstance(conn, sqlite3.Connection):
                base += " AND date(t.timestamp) >= ?"
            else:
                base += f" AND CAST(t.timestamp AS DATE) >= {ph}"
            params.append(start_date)
        if end_date:
            if isinstance(conn, sqlite3.Connection):
                base += " AND date(t.timestamp) <= ?"
            else:
                base += f" AND CAST(t.timestamp AS DATE) <= {ph}"
            params.append(end_date)
            
        base += " ORDER BY t.timestamp DESC"
        cursor.execute(base, params)
        rows = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Meus Registros"
        ws.append(["Matricula", "Nome", "Tipo", "Data/Hora", "Bairro", "Cidade"])
        for row in rows:
            ts = rf(row, 'timestamp')
            ws.append([
                rf(row, 'matricula'),
                rf(row, 'name'),
                rf(row, 'record_type'),
                ts,
                rf(row, 'neighborhood'),
                rf(row, 'city')
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name="meus_registros.xlsx", as_attachment=True)
    finally:
        try:
            conn.close()
        except:
            pass

@app.route('/api/admin/users', methods=['GET'])
@token_required
def get_users(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cursor.execute(f"SELECT id, matricula, name, role FROM Users {nolock}")
        rows = cursor.fetchall()
        users = []
        for r in rows:
            users.append({
                'id': rf(r, 'id'),
                'matricula': rf(r, 'matricula'),
                'name': rf(r, 'name'),
                'role': rf(r, 'role')
            })
        return jsonify(users)
    finally:
        try:
            conn.close()
        except:
            pass

@app.route('/api/admin/users', methods=['POST'])
@token_required
def create_user_admin(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    matricula = data.get('matricula')
    name = data.get('name')
    password_raw = data.get('password')
    new_role = data.get('role', 'user')
    if not matricula or not name or not password_raw:
        return jsonify({'message': 'Dados obrigatórios faltando'}), 400
    
    hashed = bcrypt.hashpw(password_raw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        query = f"INSERT INTO Users (matricula, password, name, role) VALUES ({ph}, {ph}, {ph}, {ph})"
        cursor.execute(query, (matricula, hashed, name, new_role))
        if isinstance(conn, sqlite3.Connection):
            conn.commit()
        elif ph == '?': 
            conn.commit()
        # mirror locally
        try:
            sconn = sqlite3.connect(sqlite_path)
            scur = sconn.cursor()
            ensure_sqlite_schema(sconn)
            scur.execute("INSERT OR REPLACE INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)", (matricula, hashed, name, new_role))
            sconn.commit()
            sconn.close()
        except: pass
        return jsonify({'message': 'Usuário criado'}), 201
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matrícula já existe'}), 409
        return jsonify({'message': msg}), 500
    finally:
        try:
            conn.close()
        except: pass

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@token_required
def update_user(curr_user_mat, role, user_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    conn = get_db_connection()
    ph = get_ph(conn)
    fields = []
    values = []
    hashed = None
    if 'matricula' in data and data['matricula']:
        fields.append(f'matricula = {ph}')
        values.append(data['matricula'])
    if 'name' in data and data['name']:
        fields.append(f'name = {ph}')
        values.append(data['name'])
    if 'role' in data and data['role']:
        fields.append(f'role = {ph}')
        values.append(data['role'])
    if 'password' in data and data['password']:
        hashed = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        fields.append(f'password = {ph}')
        values.append(hashed)
    
    if not fields:
        return jsonify({'message': 'Nada para atualizar'}), 400
        
    try:
        cursor = conn.cursor()
        # Fetch old matricula first for local mirror update
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id = {ph}", (user_id,))
        old_row = cursor.fetchone()
        old_mat = rf(old_row, 'matricula') if old_row else None

        query = f"UPDATE Users SET {', '.join(fields)} WHERE id = {ph}"
        values.append(user_id)
        cursor.execute(query, tuple(values))
        if isinstance(conn, sqlite3.Connection) or ph == '?':
            conn.commit()
            
        # Mirror update locally using old_mat
        if old_mat:
            try:
                sconn = sqlite3.connect(sqlite_path)
                ensure_sqlite_schema(sconn)
                scur = sconn.cursor()
                lfields = []
                lvals = []
                if 'matricula' in data: lfields.append("matricula = ?"); lvals.append(data['matricula'])
                if 'name' in data: lfields.append("name = ?"); lvals.append(data['name'])
                if 'role' in data: lfields.append("role = ?"); lvals.append(data['role'])
                if hashed: lfields.append("password = ?"); lvals.append(hashed)
                if lfields:
                    lvals.append(old_mat)
                    scur.execute(f"UPDATE Users SET {', '.join(lfields)} WHERE matricula = ?", tuple(lvals))
                    sconn.commit()
                sconn.close()
            except: pass
            
        return jsonify({'message': 'Usuário atualizado'}), 200
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matrícula já existe'}), 409
        return jsonify({'message': msg}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@token_required
def delete_user(curr_user_mat, role, user_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        # Fetch matricula before delete
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id = {ph}", (user_id,))
        row = cursor.fetchone()
        mat = rf(row, 'matricula') if row else None
        
        cursor.execute(f"DELETE FROM TimeRecords WHERE user_id = {ph}", (user_id,))
        cursor.execute(f"DELETE FROM Users WHERE id = {ph}", (user_id,))
        if is_sqlite or ph == '?': conn.commit()
        
        if mat:
            try:
                sconn = sqlite3.connect(sqlite_path)
                scur = sconn.cursor()
                scur.execute("DELETE FROM Users WHERE matricula = ?", (mat,))
                sconn.commit()
                sconn.close()
            except: pass
        return jsonify({'message': 'Usuário excluído'}), 200
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/users/bulk-delete', methods=['POST'])
@token_required
def bulk_delete_users(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    ids = data.get('user_ids', [])
    if not ids: return jsonify({'message': 'Nenhum selecionado'}), 400
    
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        placeholders = ', '.join([ph]*len(ids))
        
        # Get matriculas for local delete
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id IN ({placeholders})", tuple(ids))
        mats = [rf(r, 'matricula') for r in cursor.fetchall()]
        
        cursor.execute(f"DELETE FROM TimeRecords WHERE user_id IN ({placeholders})", tuple(ids))
        cursor.execute(f"DELETE FROM Users WHERE id IN ({placeholders})", tuple(ids))
        if is_sqlite or ph == '?': conn.commit()
        
        if mats:
            try:
                sconn = sqlite3.connect(sqlite_path)
                scur = sconn.cursor()
                m_ph = ', '.join(['?']*len(mats))
                scur.execute(f"DELETE FROM Users WHERE matricula IN ({m_ph})", tuple(mats))
                sconn.commit()
                sconn.close()
            except: pass
        return jsonify({'message': f'{len(ids)} excluídos'}), 200
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/report', methods=['GET'])
@token_required
def get_admin_report_excel(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    target_user_id = request.args.get('user_id')
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        query = """
            SELECT t.matricula, t.user_name AS name,
                   t.record_type, t.timestamp, t.neighborhood, t.city
            FROM TimeRecords t
        """
        params = []
        if target_user_id:
            query += f" WHERE t.user_id = {ph}"
            params.append(target_user_id)
        query += " ORDER BY t.timestamp DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        wb = openpyxl.Workbook()
        if target_user_id:
            ws = wb.active
            ws.title = "Relatorio"
            ws.append(["Matricula", "Nome", "Tipo", "Data/Hora", "Bairro", "Cidade"])
            for r in rows:
                ws.append([rf(r,'matricula'), rf(r,'name'), rf(r,'record_type'), rf(r,'timestamp'), rf(r,'neighborhood'), rf(r,'city')])
        else:
            wb.remove(wb.active)
            groups = {}
            for r in rows:
                k = (rf(r,'matricula'), rf(r,'name'))
                groups.setdefault(k, []).append(r)
            for (m, n), items in groups.items():
                ws = wb.create_sheet(title=(n or m or "User")[:30])
                ws.append(["Matricula", "Nome", "Tipo", "Data/Hora", "Bairro", "Cidade"])
                for r in items:
                    ws.append([rf(r,'matricula'), rf(r,'name'), rf(r,'record_type'), rf(r,'timestamp'), rf(r,'neighborhood'), rf(r,'city')])
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, download_name="relatorio_admin.xlsx", as_attachment=True)
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/export', methods=['GET'])
@token_required
def export_excel_legacy(curr_user_mat, role):
    return get_admin_report_excel(curr_user_mat, role)

@app.route('/api/admin/sync_all', methods=['POST'])
@token_required
def sync_all_users_admin(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    refresh_local_users()
    return jsonify({'message': 'Sincronização de usuários solicitada'}), 200

def refresh_local_users():
    """Helper to pull users from SQL into local SQLite."""
    try:
        conn = get_db_connection()
        if isinstance(conn, sqlite3.Connection):
            return # Already strictly local
        cursor = conn.cursor()
        cursor.execute("SELECT matricula, password, name, role FROM Users WITH (NOLOCK)")
        users = cursor.fetchall()
        
        sconn = sqlite3.connect(sqlite_path)
        scur = sconn.cursor()
        ensure_sqlite_schema(sconn)
        for u in users:
            scur.execute("INSERT OR REPLACE INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)", 
                         (rf(u,'matricula'), rf(u,'password'), rf(u,'name'), rf(u,'role')))
        sconn.commit()
        sconn.close()
        conn.close()
    except: pass

@app.route('/api/sync', methods=['POST'])
@token_required
def sync_now(curr_user_mat, role):
    migrated, errs = perform_sync_for_user(curr_user_mat)
    return jsonify({'message': f'Sincronização concluída. {migrated} registros enviados.', 'migrated': migrated, 'errors': errs}), 200

def perform_sync_for_user(user_matricula):
    """
    Core sync logic that can be called via API or background thread.
    Returns (migrated_count, errors_list)
    """
    forced = os.getenv('FORCE_ONLINE', 'false').lower() == 'true'
    is_sql = sql_online() or forced
    
    # Initialize user info variables
    sql_user_id = None
    user_name = None
    local_user_id = None
    l_user_name = None
    sync_user_id = None

    # Get user info from SQL Server if online
    if is_sql:
        try:
            conn_sql = get_db_connection()
            sql_user_id, user_name = get_user_info_by_matricula(user_matricula, conn_sql)
        except: pass
        finally:
            try:
                if conn_sql: conn_sql.close()
            except: pass

    # Get user info from local SQLite
    lconn = sqlite3.connect(sqlite_path)
    lconn.row_factory = sqlite3.Row
    try:
        local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
    finally:
        try: lconn.close()
        except: pass

    if not user_name:
         user_name = l_user_name
    
    sync_user_id = sql_user_id if sql_user_id else local_user_id
    
    if not is_sql:
        # Local-only sync (SQL to SQLite Mirror)
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            # Catch matricula, null matricula, or empty string matricula
            scur.execute("SELECT id, record_type, neighborhood, city, timestamp FROM OfflineQueue WHERE matricula = ? OR matricula IS NULL OR matricula = '' AND user_id = ? ORDER BY timestamp ASC", (user_matricula, local_user_id))
            rows = scur.fetchall()
            migrated = 0
            for r in rows:
                scur.execute(
                    "INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, neighborhood, city, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (sync_user_id, user_matricula, user_name, rf(r, 'record_type'), rf(r, 'neighborhood'), rf(r, 'city'), rf(r, 'timestamp'))
                )
                scur.execute("DELETE FROM OfflineQueue WHERE id = ?", (rf(r, 'id'),))
                migrated += 1
            sconn.commit()
            sconn.close()
            return migrated, []
        except Exception as e:
            return 0, [str(e)]

    # SQL Server Sync
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        sql_user_id_remote, user_name_remote = get_user_info_by_matricula(user_matricula, conn)
        if user_name_remote:
            user_name = user_name_remote
            sync_user_id = sql_user_id_remote

        existing_sigs = set()
        errs = []
        try:
            import pymssql
            sph = '%s' if isinstance(conn, pymssql.Connection) else '?'
            cursor.execute(f"SELECT record_type, timestamp FROM TimeRecords WITH (NOLOCK) WHERE matricula = {sph}", (user_matricula,))
            for r in cursor.fetchall():
                ts_str = str(rf(r, 1)) # Use index if not as_dict
                if '.' in ts_str: ts_str = ts_str.split('.')[0]
                existing_sigs.add((rf(r, 0), ts_str))
        except Exception as e:
            errs.append(f"Error fetching online records: {e}")

        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        ensure_sqlite_schema(sconn)
        scur = sconn.cursor()
        
        migrated = 0
        # 1. Sync local TimeRecords
        try:
            scur.execute("SELECT record_type, timestamp, neighborhood, city, matricula FROM TimeRecords WHERE matricula = ? OR ((matricula IS NULL OR matricula = '') AND user_id = ?)", (user_matricula, local_user_id))
            local_rows = scur.fetchall()
            for r in local_rows:
                ts_val = rf(r, 'timestamp')
                # Standardize to datetime object for SQL Server
                if isinstance(ts_val, str):
                    try:
                        if '.' in ts_val: ts_val = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S.%f")
                        else: ts_val = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S")
                    except: pass

                cmp_ts = str(ts_val).split('.')[0]
                if (rf(r, 'record_type'), cmp_ts) not in existing_sigs:
                    try:
                        import pymssql
                        sph = '%s' if isinstance(conn, pymssql.Connection) else '?'
                        query_ins = f"INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, neighborhood, city, timestamp) VALUES ({sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph})"
                        cursor.execute(query_ins, (sync_user_id, user_matricula, user_name, rf(r, 'record_type'), rf(r, 'neighborhood'), rf(r, 'city'), ts_val))
                        migrated += 1
                        existing_sigs.add((rf(r, 'record_type'), cmp_ts))
                        # Heal local
                        if not rf(r, 'matricula'):
                             scur.execute("UPDATE TimeRecords SET matricula = ?, user_name = ? WHERE timestamp = ? AND record_type = ?", (user_matricula, user_name, rf(r, 'timestamp'), rf(r, 'record_type')))
                    except Exception as e:
                        errs.append(str(e))
        except: pass

        # 2. Process OfflineQueue
        scur.execute("SELECT id, record_type, neighborhood, city, timestamp, user_id, matricula FROM OfflineQueue WHERE matricula = ? OR ((matricula IS NULL OR matricula = '') AND (user_id = ? OR user_id = ?)) ORDER BY timestamp ASC", (user_matricula, local_user_id, sql_user_id))
        rows = scur.fetchall()
        for r in rows:
            ts_val = rf(r, 'timestamp')
            # Standardize for SQL Server
            ts_dt = ts_val
            if isinstance(ts_val, str):
                try:
                    if '.' in ts_val: ts_dt = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S.%f")
                    else: ts_dt = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S")
                except: pass

            cmp_ts = str(ts_dt).split('.')[0]
            if (rf(r, 'record_type'), cmp_ts) not in existing_sigs:
                try:
                    import pymssql
                    sph = '%s' if isinstance(conn, pymssql.Connection) else '?'
                    cursor.execute(f"INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, neighborhood, city, timestamp) VALUES ({sph},{sph},{sph},{sph},{sph},{sph},{sph})",
                                   (sync_user_id, user_matricula, user_name, rf(r, 'record_type'), rf(r, 'neighborhood'), rf(r, 'city'), ts_dt))
                    migrated += 1
                    scur.execute("DELETE FROM OfflineQueue WHERE id = ?", (rf(r, 'id'),))
                except Exception as e:
                    errs.append(str(e))
            else:
                scur.execute("DELETE FROM OfflineQueue WHERE id = ?", (rf(r, 'id'),))

        sconn.commit()
        sconn.close()
        refresh_local_users()
        return migrated, errs
    finally:
        try: conn.close()
        except: pass

def auto_sync_all():
    """Finds all users with pending items and syncs them."""
    print("DEBUG: Starting automatic background synchronization...")
    try:
        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        scur = sconn.cursor()
        # Find all distinct matriculas and user_ids in the queue
        scur.execute("SELECT DISTINCT matricula, user_id FROM OfflineQueue")
        rows = scur.fetchall()
        
        synced_users = set()
        total_migrated = 0
        
        for r in rows:
            m = rf(r, 'matricula')
            uid = rf(r, 'user_id')
            
            # If matricula is missing, try to find it via user_id
            if not m or m == '':
                scur.execute("SELECT matricula FROM Users WHERE id = ? OR matricula = ?", (uid, str(uid)))
                u_row = scur.fetchone()
                m = rf(u_row, 'matricula') if u_row else None
            
            if m and m not in synced_users:
                print(f"DEBUG: Auto-syncing for matricula: {m}")
                migrated, errs = perform_sync_for_user(m)
                total_migrated += migrated
                synced_users.add(m)
        
        sconn.close()
        if total_migrated > 0:
            print(f"DEBUG: Automatic sync complete. {total_migrated} records synchronized.")
    except Exception as e:
        print(f"DEBUG: Auto-sync error: {e}")

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5005'))
    ensure_default_admin()
    migrate_local_data()
    start_health_check()
    app.run(host='0.0.0.0', debug=False, port=port)
